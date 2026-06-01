import io, re, time, random, requests, pickle, gc
import pandas as pd
import streamlit as st
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="사업장 주소·전화번호 보완", page_icon="📋", layout="centered")
st.markdown("""<style>
#MainMenu,footer,header{visibility:hidden}
.block-container{max-width:660px!important;padding:2rem 1rem!important}
div.stButton>button{background:#3b82f6!important;color:#fff!important;border:none!important;border-radius:8px!important;font-weight:600!important;width:100%!important}
div.stDownloadButton>button{background:#10b981!important;color:#fff!important;border:none!important;border-radius:8px!important;font-weight:600!important;width:100%!important}
</style>""", unsafe_allow_html=True)

FSC_KEY   = "2446da601452d231f1ff77a4c5ac9598896028fb96e8be586bf7f4be58b6231c"
KAKAO_KEY = "e8eef7246a2cd0f3d2d55af8a806e3ef"
NAV_HDR   = {"User-Agent":"Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15","Referer":"https://map.naver.com/","Accept-Language":"ko-KR,ko;q=0.9"}
WORKERS   = 20

def kakao_keyword(name, addr):
    try:
        r = requests.get("https://dapi.kakao.com/v2/local/search/keyword.json",
            headers={"Authorization":f"KakaoAK {KAKAO_KEY}"},
            params={"query":f"{name} {addr[:20]}","size":5}, timeout=5)
        docs = r.json().get("documents",[])
        m = next((d for d in docs if name[:4] in d.get("place_name","")), None)
        return m or (docs[0] if docs else None)
    except: return None

def fsc_info(name, biz_no):
    try:
        r = requests.get(
            "https://apis.data.go.kr/1160100/service/GetCorpBasicInfoService_V2/getCorpOutline_V2",
            params={"serviceKey":FSC_KEY,"pageNo":1,"numOfRows":5,"resultType":"json","corpNm":name}, timeout=5)
        items = r.json().get("response",{}).get("body",{}).get("items",{}).get("item",[])
        if not items: return None
        if isinstance(items, dict): items = [items]
        bc = re.sub(r'\D','',str(biz_no))
        m = next((i for i in items if re.sub(r'\D','',str(i.get("bzno",""))) == bc), items[0])
        return {"phone":m.get("enpTlno",""), "addr":m.get("enpBsadr","") or m.get("enpDtadr","")}
    except: return None

def naver_phone(name, addr):
    clean = re.sub(r'주식회사|유한회사|\(주\)|\(유\)|[(){}\[\]]','',name).strip()
    gu = re.search(r'[\w]+구', addr); gu = gu.group() if gu else ""
    try:
        r = requests.get("https://m.map.naver.com/search2/search.nhn",
            headers=NAV_HDR, params={"query":f"{clean} {gu}","type":"all"}, timeout=8)
        if r.status_code == 429: time.sleep(30); return naver_phone(name, addr)
        for pat in [r'tel:(0\d{1,2}-?\d{3,4}-?\d{4})',r'(?<!\d)(0\d{1,2}-\d{3,4}-\d{4})(?!\d)']:
            f = re.findall(pat, r.text)
            if f: return f[0]
    except: pass
    return ""

def process_one(row_data):
    idx, name, addr, biz_no, is_corp = row_data
    phone=""; full_addr=addr; lat=None; lon=None
    if is_corp:
        info = fsc_info(name, biz_no)
        if info:
            phone = info.get("phone","")
            if info.get("addr"): full_addr = info["addr"]
        time.sleep(0.1)
    kk = kakao_keyword(name, addr)
    if kk:
        if not phone: phone = kk.get("phone","")
        lat = float(kk.get("y") or 0) or None
        lon = float(kk.get("x") or 0) or None
        road = kk.get("road_address_name","")
        if road: full_addr = road
    if not phone: phone = naver_phone(name, full_addr)
    return idx, {"phone":phone, "addr":full_addr, "lat":lat, "lon":lon}

def make_excel(df):
    wb = Workbook(); ws = wb.active; ws.title = "보완완료"
    def thin():
        s = Side(style="thin",color="D0D0D0")
        return Border(left=s,right=s,top=s,bottom=s)
    hdrs = list(df.columns)
    widths = {"사업장명":26,"업종":18,"법인여부":8,"주소":42,"전화번호":14,"가입자수":8,"당월고지금액":14,"위도":12,"경도":12}
    CENTER = {"법인여부","가입자수","당월고지금액","위도","경도","전화번호"}
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.fill = PatternFill("solid",start_color="1F3864")
        cell.font = Font(name="Arial",size=10,bold=True,color="FFFFFF")
        cell.alignment = Alignment(horizontal="center",vertical="center")
        cell.border = thin()
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h,12)
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"
    for ri,row in enumerate(df.itertuples(index=False),2):
        for ci,(h,v) in enumerate(zip(hdrs,row),1):
            if pd.isna(v) or str(v) in ("nan","None"): v=""
            cell = ws.cell(ri,ci,v)
            cell.font = Font(name="Arial",size=10)
            cell.border = thin()
            cell.fill = PatternFill("solid",start_color="FFFFFF" if ri%2==0 else "F8FAFC")
            cell.alignment = Alignment(horizontal="center" if h in CENTER else "left",vertical="center")
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── UI ──────────────────────────────────────────────────
st.title("📋 사업장 주소·전화번호 보완")
st.caption("대시보드에서 다운로드한 지역 파일을 업로드하세요")
st.divider()

uploaded = st.file_uploader("파일 선택 (국민연금_XX시_XX구_사업장.xlsx)", type=["xlsx"])

if uploaded:
    # 필요한 컬럼만 읽기 (메모리 절약)
    df = pd.read_excel(uploaded)
    df.columns = [c.strip() for c in df.columns]
    gc.collect()  # 메모리 정리

    name_col = next((c for c in df.columns if "사업장명" in c or "상호" in c), None)
    addr_col = next((c for c in df.columns if "주소" in c or "도로명" in c), None)
    biz_col  = next((c for c in df.columns if "사업자" in c and "번호" in c), None)
    type_col = next((c for c in df.columns if "형태" in c or "법인" in c), None)
    emp_col  = next((c for c in df.columns if "가입자" in c), None)
    pay_col  = next((c for c in df.columns if "고지" in c), None)
    ind_col  = next((c for c in df.columns if "업종" in c), None)

    for col in ["전화번호"]:
        if col not in df.columns: df[col] = ""
        else: df[col] = df[col].fillna("").astype(str).replace("nan","")
    for col in ["위도","경도"]:
        if col not in df.columns: df[col] = None
        else: df[col] = pd.to_numeric(df[col], errors="coerce")

    total = len(df)
    fname = uploaded.name.replace(".xlsx","")
    corp  = int((df[type_col].astype(str).str.contains("법인|^1$")).sum()) if type_col else 0

    st.success(f"✅ **{fname}** — {total:,}개 사업장 (법인 {corp:,}개)")
    

    CHUNK = 300

    # 세션 초기화 (파일 바뀌면 리셋)
    if "df_work" not in st.session_state or st.session_state.get("work_fname") != fname:
        st.session_state["df_work"]    = df.copy()
        st.session_state["work_fname"] = fname
        st.session_state["work_done"]  = 0
        st.session_state["phone_ok"]   = 0
        st.session_state["coord_ok"]   = 0
        st.session_state["finished"]   = False

    wdf      = st.session_state["df_work"]
    done_cnt = st.session_state["work_done"]
    finished = st.session_state["finished"]

    # 진행바
    pct = done_cnt / total if total > 0 else 0
    st.progress(pct, text=f"{done_cnt:,}/{total:,}개 완료 | 📞{st.session_state['phone_ok']} 📍{st.session_state['coord_ok']}")

    def get_out(wdf):
        col_map = {}
        if name_col: col_map[name_col]="사업장명"
        if type_col: col_map[type_col]="법인여부"
        if addr_col: col_map[addr_col]="주소"
        col_map["전화번호"]="전화번호"
        if emp_col:  col_map[emp_col]="가입자수"
        if pay_col:  col_map[pay_col]="당월고지금액"
        if ind_col:  col_map[ind_col]="업종"
        col_map["위도"]="위도"; col_map["경도"]="경도"
        out = wdf[[c for c in col_map if c in wdf.columns]].rename(columns=col_map)
        return out[out["전화번호"].notna() & (out["전화번호"]!="") &
                   out["주소"].notna()     & (out["주소"]!="")].reset_index(drop=True)

    if finished:
        out = get_out(wdf)
        st.success(f"✅ 완료! 📞 {st.session_state['phone_ok']:,}개 | 📍 {st.session_state['coord_ok']:,}개 | 최종 {len(out):,}개")
        st.balloons()
        st.download_button("⬇️ 보완 완료 엑셀 다운로드",
            make_excel(out), fname+"_보완완료.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

    elif done_cnt >= total:
        st.session_state["finished"] = True
        st.rerun()

    else:
        btn_label = "🚀 보완 시작" if done_cnt == 0 else f"▶️ 계속 ({done_cnt:,}/{total:,})"
        if st.button(btn_label, type="primary", use_container_width=True):
            chunk_idx = list(wdf.index[done_cnt:done_cnt+CHUNK])
            tasks = []
            for idx in chunk_idx:
                row = wdf.loc[idx]
                tasks.append((idx,
                    str(row.get(name_col,"")) if name_col else "",
                    str(row.get(addr_col,"")) if addr_col else "",
                    str(row.get(biz_col,""))  if biz_col  else "",
                    str(row.get(type_col,"")) in ["법인","1"] if type_col else False))

            prog = st.progress(0, text=f"처리 중... (0/{len(tasks)})")
            completed = 0

            with ThreadPoolExecutor(max_workers=WORKERS) as executor:
                futures = {executor.submit(process_one, t): t[0] for t in tasks}
                for future in as_completed(futures):
                    try:
                        idx, res = future.result()
                        wdf.loc[idx,"전화번호"] = str(res["phone"])
                        wdf.loc[idx,"위도"]    = res["lat"]
                        wdf.loc[idx,"경도"]    = res["lon"]
                        if addr_col and res["addr"]: wdf.loc[idx,addr_col] = res["addr"]
                        if res["phone"]: st.session_state["phone_ok"] += 1
                        if res["lat"]:   st.session_state["coord_ok"] += 1
                    except: pass
                    completed += 1
                    prog.progress(completed/len(tasks), text=f"처리 중... ({completed}/{len(tasks)})")

            st.session_state["df_work"]   = wdf
            st.session_state["work_done"] = done_cnt + len(chunk_idx)
            prog.empty()
            st.rerun()

        if done_cnt > 0:
            st.caption(f"새로고침 되어도 {done_cnt:,}개 저장됨. 버튼 눌러 계속하세요.")
